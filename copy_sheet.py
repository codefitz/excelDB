# importing openpyxl module 
import openpyxl as xl
import sys

# Check if the correct number of command-line arguments are provided
if len(sys.argv) != 3:
    print("Usage: python script.py <source_filename> <destination_filename>")
    sys.exit(1)

# Get the filenames from the command-line arguments
filename = sys.argv[1]
filename1 = sys.argv[2]

# opening the source excel file 
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file 
wb2 = xl.load_workbook(filename1)

# Get the sheet name
source_sheet_name = ws1.title

# Check if a sheet with the same name exists in the destination workbook
if source_sheet_name in wb2.sheetnames:
    # If it exists, find a unique name for the copy
    copy_suffix = 1
    new_sheet_name = f"{source_sheet_name}_copy{copy_suffix}"
    while new_sheet_name in wb2.sheetnames:
        copy_suffix += 1
        new_sheet_name = f"{source_sheet_name}_copy{copy_suffix}"
else:
    new_sheet_name = source_sheet_name

# Create a new sheet in the destination workbook with the unique name
ws2 = wb2.create_sheet(title=new_sheet_name)

# calculate total number of rows and 
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source 
# excel file to destination excel file 
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file 
        c = ws1.cell(row=i, column=j)
        
        # writing the read value to destination excel file 
        ws2.cell(row=i, column=j).value = c.value 

# saving the destination excel file 
wb2.save(filename1)
